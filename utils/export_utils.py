from io import BytesIO
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _join(value):
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    return str(value or "")


def create_excel(reviewed_result):
    rows = []
    for tc in reviewed_result.get("test_cases", []):
        steps = tc.get("steps", [])
        rows.append({
            "Test Case ID": tc.get("test_case_id", ""),
            "Requirement ID": tc.get("requirement_id", ""),
            "Title": tc.get("title", ""), "Module": tc.get("module", ""),
            "Category": tc.get("category", ""), "Priority": tc.get("priority", ""),
            "Objective": tc.get("objective", ""), "User Role": tc.get("user_role", ""),
            "Preconditions": _join(tc.get("preconditions", [])), "Test Data": _join(tc.get("test_data", [])),
            "Test Steps": "\n".join(f'{s.get("step_number")}. {s.get("action", "")}' for s in steps),
            "Expected Results": "\n".join(f'{s.get("step_number")}. {s.get("expected_result", "")}' for s in steps),
            "Final Expected Result": tc.get("final_expected_result", ""),
            "Postconditions": _join(tc.get("postconditions", [])),
            "Automation Candidate": tc.get("automation_candidate", ""),
            "Traceability": ", ".join(tc.get("traceability", [])),
            "Assumptions": _join(tc.get("assumptions", [])),
            "Actual Result": tc.get("actual_result", ""), "Status": tc.get("status", "Not Executed")
        })
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name="Test Cases", index=False)
        pd.DataFrame(reviewed_result.get("traceability_matrix", [])).to_excel(writer, sheet_name="Traceability", index=False)
        pd.DataFrame([reviewed_result.get("coverage_summary", {})]).to_excel(writer, sheet_name="Coverage Summary", index=False)
        for ws in writer.book.worksheets:
            fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                max_len = max((len(str(c.value or "")) for c in col), default=12)
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 50)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    return output.getvalue()
